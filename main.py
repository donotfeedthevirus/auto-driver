import openpyxl
import time, random, datetime
from autodriver_robot import AutoDriverRobot

def split_list(data, chunk_size=20):
    for i in range(0, len(data), chunk_size):
        yield data[i:i + chunk_size]

def process_chunk(cnh_list, p1_sheet, f1_sheet, username, password):
    robot = AutoDriverRobot(username, password)
    robot.login()

    for cnh in cnh_list:
        cnh = str(cnh).zfill(11)

        try:
            name, cpf = robot.search_by_cnh(cnh)
            print(f"CNH {cnh} Coletou Nome: {name}, CPF: {cpf}")
        except Exception as e:
            print(f"Erro ao coletar Nome e CPF da CNH {cnh}: {e}")
            f1_sheet.append([cnh])
            time.sleep(60)
            continue

        try:
            phone = robot.search_by_cpf(cpf)
            print(f"CNH {cnh} Coletou Telefone: {phone}")
        except Exception as e:
            print(f"Erro ao coletar Telefone da CNH {cnh}: {e}")
            f1_sheet.append([cnh])
            time.sleep(60)
            continue

        p1_sheet.append([cnh, name, phone])
        print(f"CNH {cnh} Processada - Nome: {name}, Telefone: {phone}")
        time.sleep(random.randint(30, 120))

    robot.close()

def main():
    date_str = datetime.datetime.now().strftime("%d-%m-%Y")
    p1_filename = f"p1-{date_str}.xlsx"
    f1_filename = f"f1-{date_str}.xlsx"

    p1_wb = openpyxl.Workbook()
    p1_sheet = p1_wb.active
    p1_sheet.title = "CNHs Processadas"
    p1_sheet.append(["CNH", "NOME", "TELEFONE"])

    f1_wb = openpyxl.Workbook()
    f1_sheet = f1_wb.active
    f1_sheet.title = "CNHs Não Processadas"
    f1_sheet.append(["CNH"])
    
    input_sheet_location = "./entrada.xlsx"
    input_wb = openpyxl.load_workbook(input_sheet_location)
    input_sheet = input_wb.active

    cnh_list = [input_sheet.cell(row=i, column=1).value for i in range(2, input_sheet.max_row + 1)]
    cnh_list = [c for c in cnh_list if c]

    chunks = list(split_list(cnh_list, 20))

    username = "canever"
    password = "102030"

    for chunk in chunks:
        process_chunk(chunk, p1_sheet, f1_sheet, username, password)

        p1_wb.save(p1_filename)
        f1_wb.save(f1_filename)

    p1_wb.save(p1_filename)
    f1_wb.save(f1_filename)
    print(f"Processamento finalizado. CNHs processadas: {p1_filename}, Not fetched file: {f1_filename}")

if __name__ == '__main__':
    main()